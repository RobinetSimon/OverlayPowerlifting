import json
from openpyxl import load_workbook

# Load the workbook with data_only=True to read calculated values
wb = load_workbook("../dataset/Régional FA JEUNES COMP.xlsm", data_only=True)
ws = wb.active

# Column indices
fields_index = {
    "club": 1,
    "sex": 2,
    "category_age": 4,
    "last_name": 5,
    "first_name": 6,
    "weight_cat": 8,
    "squat": [11, 12, 13],
    "bench": [14, 15, 16],
    "deadlift": [17, 18, 19],
    "total": 20
}

def get_attempt_status(cell):
    fill = cell.fill
    strike = cell.font.strike

    # Couleur RGB (au format ARGB 8 chiffres) — '0000000B' → 11, '0000001F' → 31
    color = fill.start_color.index if fill.start_color else None

    # Traitement des codes connus
    if color in ('0000000B', 11):  # vert / valide
        return "valid"
    elif color in ('0000001F', 31) and strike:  # violet + barré = échec
        return "invalid"
    elif color in ('00000000', None):
        return "pending"
    else:
        return "pending"

athletes = []

# Iterate through rows starting from row 8
for row in ws.iter_rows(min_row=8):
    if not row[fields_index["last_name"]].value or not row[fields_index["first_name"]].value:
        continue

    athlete = {
        "club": row[fields_index["club"]].value,
        "sex": row[fields_index["sex"]].value,
        "category_age": row[fields_index["category_age"]].value,
        "last_name": row[fields_index["last_name"]].value,
        "first_name": row[fields_index["first_name"]].value,
        "weight_category": row[fields_index["weight_cat"]].value,
        "attempts": {
            "squat": [],
            "bench_press": [],
            "deadlift": []
        },
        "total": row[fields_index["total"]].value
    }

    for col in fields_index["squat"]:
        cell = row[col]
        athlete["attempts"]["squat"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    for col in fields_index["bench"]:
        cell = row[col]
        athlete["attempts"]["bench_press"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    for col in fields_index["deadlift"]:
        cell = row[col]
        athlete["attempts"]["deadlift"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    athletes.append(athlete)

# Export the results to JSON
output_path = "../overlay-powerlifting/public/json/datas.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(athletes, f, indent=2, ensure_ascii=False)

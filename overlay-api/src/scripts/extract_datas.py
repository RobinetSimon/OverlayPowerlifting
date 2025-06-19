import json
import sys
from openpyxl import load_workbook

if len(sys.argv) < 3:
    print("Usage: python script.py <chemin_excel> <chemin_json>")
    sys.exit(1)

chemin_excel = sys.argv[1]
chemin_json = sys.argv[2]

# Load the workbook with data_only=True to read calculated values
wb = load_workbook(chemin_excel, data_only=True)
ws = wb.active

# Column indices
fields_index = {
    "club": 1,
    "sex": 2,
    "category_age": 4,
    "last_name": 5,
    "first_name": 6,
    "weight_cat": 8,
    "squat": [11, 12, 13],
    "bench": [14, 15, 16],
    "deadlift": [17, 18, 19],
    "total": 20
}

def get_attempt_status(cell):
    fill = cell.fill
    strike = cell.font.strike

    color = fill.start_color.index if fill.start_color else None

    if color in ('0000000B', 11):  # vert / valide
        return "valid"
    elif color in ('0000001F', 31) and strike:  # violet + barré = échec
        return "invalid"
    elif color in ('00000000', None):
        return "pending"
    else:
        return "pending"

athletes = []

for row in ws.iter_rows(min_row=8):
    if not row[fields_index["last_name"]].value or not row[fields_index["first_name"]].value:
        continue

    athlete = {
        "club": row[fields_index["club"]].value,
        "sex": row[fields_index["sex"]].value,
        "category_age": row[fields_index["category_age"]].value,
        "last_name": row[fields_index["last_name"]].value,
        "first_name": row[fields_index["first_name"]].value,
        "weight_category": row[fields_index["weight_cat"]].value,
        "attempts": {
            "squat": [],
            "bench_press": [],
            "deadlift": []
        },
        "total": row[fields_index["total"]].value
    }

    for col in fields_index["squat"]:
        cell = row[col]
        athlete["attempts"]["squat"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    for col in fields_index["bench"]:
        cell = row[col]
        athlete["attempts"]["bench_press"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    for col in fields_index["deadlift"]:
        cell = row[col]
        athlete["attempts"]["deadlift"].append({
            "weight": cell.value,
            "status": get_attempt_status(cell)
        })

    athletes.append(athlete)

# Export the results to JSON
with open(chemin_json, "w", encoding="utf-8") as f:
    json.dump(athletes, f, indent=2, ensure_ascii=False)
